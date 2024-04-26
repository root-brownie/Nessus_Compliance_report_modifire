import openpyxl

# Load the Excel file
import openpyxl

#input file 
file_path = input("Enter the file path: ")
workbook = openpyxl.load_workbook(file_path)

# Select the active sheet
sheet = workbook.active

# Define the columns to keep (N, P, Q, R)
columns_to_keep = ['N', 'P', 'Q', 'R']

# Get the maximum column index
max_column_index = sheet.max_column

# Get the maximum row index
max_row = sheet.max_row

# Iterate over the columns in reverse order
for column in reversed(range(1, max_column_index + 1)):
    column_letter = openpyxl.utils.get_column_letter(column)
    
    # Check if the column should be deleted
    if column_letter not in columns_to_keep:
        sheet.delete_cols(column)

max_row = sheet.max_row

# Insert a blank column between Column A and Column B
sheet.insert_cols(2)

# Move the content of Column E to the newly created column (Column B)
for row in range(1, max_row + 1):
    value = sheet.cell(row=row, column=5).value  # Get value from Column E
    sheet.cell(row=row, column=2).value = value  # Assign value to Column B

# Delete Column E After moving
sheet.delete_cols(5)

# Loop through each row in column K
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=4, max_col=4):
    cell = row[0]
    if cell.value and 'Impact' in cell.value:
        # Delete everything before "Impact"
        cell.value = cell.value.split('Impact', 1)[-1]
       
       #delete extra spaces apfter impact
        cell.value = cell.value.split('\n', 2)[-1]

# Save the modified workbook
save_path = input("Enter the save path: ")
workbook.save(save_path)


print("Report Genetated")
